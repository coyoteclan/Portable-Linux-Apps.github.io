#!/usr/bin/env python3
"""Convert a gettext .pot file to an XLSX spreadsheet for mass translation.

Usage:
    python3 pot2xlsx.py locales/apps-list.pot -o locales/apps-list.xlsx

Columns:
    A - App name  (from #: comment)
    B - English   (msgid)
    C - Translation (msgstr, empty)
    D - Context   (msgctxt if any)

In Google Sheets, translate column C with:
    =GOOGLETRANSLATE(B2, "en", "it")
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def unescape_po(s: str) -> str:
    """Unescape a PO string (\\n -> newline, \\" -> ", etc.)."""
    return (
        s.replace("\\n", "\n")
        .replace("\\t", "\t")
        .replace('\\"', '"')
        .replace("\\\\", "\\")
    )


def extract_po_string(block: str, keyword: str) -> str:
    """Extract a multi-line PO string value for the given keyword (msgid, msgstr, msgctxt).
    
    Handles both single-line:
        msgid "text"
    And multi-line:
        msgid ""
        "line1 "
        "line2"
    """
    # Find the keyword line
    pattern = rf'^{keyword}\s+"((?:[^"\\]|\\.)*)"'
    match = re.search(pattern, block, re.MULTILINE)
    if not match:
        return ""
    
    value = match.group(1)
    
    # If the value is empty, check for continuation lines
    if not value:
        # Find all continuation lines after the keyword
        lines = block.split("\n")
        keyword_line_idx = None
        for i, line in enumerate(lines):
            if line.strip().startswith(f'{keyword} '):
                keyword_line_idx = i
                break
        
        if keyword_line_idx is not None:
            for i in range(keyword_line_idx + 1, len(lines)):
                line = lines[i].strip()
                if line.startswith('"') and line.endswith('"'):
                    value += line[1:-1]  # strip outer quotes
                else:
                    break  # no more continuation lines
    
    return unescape_po(value)


def parse_pot(path: str) -> list[dict]:
    """Parse a .pot file into a list of {context, msgid, msgstr, app_name}.
    
    Note: A single block may have multiple #: location comments (multiple apps
    sharing the same description). We expand these into separate rows so every
    app gets its own line in the spreadsheet.
    """
    content = Path(path).read_text(encoding="utf-8")
    entries = []

    # Split into blocks by blank lines.
    blocks = re.split(r'\n\n+', content)

    for block in blocks:
        lines = block.strip().split("\n")
        if not lines:
            continue

        # Extract ALL location comments (there may be multiple).
        app_names = []
        for line in lines:
            if line.startswith("#: "):
                # e.g. "#: graphest:1 superslicer-prerelease-bin:1"
                parts = line[3:].strip()
                for loc in parts.split():
                    name = loc.split(":")[0].strip()
                    if name:
                        app_names.append(name)

        if not app_names:
            continue

        # Extract msgctxt if present.
        ctx = extract_po_string(block, "msgctxt")

        # Extract msgid.
        msgid = extract_po_string(block, "msgid")
        if not msgid:
            continue  # skip header entry

        # Extract msgstr.
        msgstr = extract_po_string(block, "msgstr")

        # Emit one row per app name so every app appears in the spreadsheet.
        for app_name in app_names:
            entries.append({
                "context": ctx,
                "msgid": msgid,
                "msgstr": msgstr,
                "app_name": app_name,
            })

    return entries


def to_xlsx(entries: list[dict], output: str, lang: str = "it"):
    """Write entries to an XLSX file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Translations"

    # Styling.
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["App Name", "English", "Translation", "Context"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows.
    for idx, entry in enumerate(entries, 2):
        ws.cell(row=idx, column=1, value=entry["app_name"]).border = thin_border
        ws.cell(row=idx, column=2, value=entry["msgid"]).border = thin_border
        # Add GOOGLETRANSLATE formula instead of empty string
        formula = f'=GOOGLETRANSLATE(B{idx},"en","{lang}")'
        ws.cell(row=idx, column=3, value=formula).border = thin_border
        ws.cell(row=idx, column=4, value=entry["context"]).border = thin_border

    # Column widths.
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 15

    # Freeze top row.
    ws.freeze_panes = "A2"

    wb.save(output)
    print(f"Wrote {len(entries)} entries to {output}")


def main():
    parser = argparse.ArgumentParser(description="Convert .pot to XLSX for Google Sheets translation")
    parser.add_argument("pot", help="Input .pot file")
    parser.add_argument("-o", "--output", help="Output .xlsx file (default: same name with .xlsx)")
    parser.add_argument("-l", "--lang", default="it", help="Target language code (default: it)")
    args = parser.parse_args()

    output = args.output or args.pot.rsplit(".", 1)[0] + ".xlsx"
    entries = parse_pot(args.pot)
    to_xlsx(entries, output, args.lang)


if __name__ == "__main__":
    main()
